import openpyxl
from openpyxl.styles.borders import Border, Side
from random import randint
import random
import subprocess
import os


wbObj = openpyxl.load_workbook('list.xlsx')
wbk = wbObj.worksheets[0]

wbObj_answers = openpyxl.load_workbook('list.xlsx')
wbk_answers = wbObj_answers.worksheets[0]

style_border = 'medium'
color_border = 'cccccc'
border = Border(left=Side(style=style_border, color=color_border),
                 right=Side(style=style_border, color=color_border),
                 top=Side(style=style_border, color=color_border),
                 bottom=Side(style=style_border, color=color_border)
                )


def calc(action, n1, n2):
    match action:
        case '+':
            return n1 + n2
        case '-':
            return n1 - n2
        case ':':
            return n1 / n2
        case '×':
            return n1 * n2


def generate_examples(n=22, types_action=['+', '-', '×', ':'], number_range=[1, 100], only_positive=False):
    examples_list = []
    result_list = []

    for _ in range(n):
        action = random.choice(types_action)

        num1 = randint(*number_range)
        num2 = randint(*number_range)

        if only_positive:
            num2 = randint(number_range[0], num1)

        num1_str, num2_str = list(str(num1)), list(str(num2))

        examples_list.append([*num1_str, action, *num2_str, '='])
        result_list.append(list(str(calc(action, num1, num2))))

    return [examples_list, result_list]


def print_examples_cell(current_wbk, examples, start_row=4, start_column=5, answers_list=[]):

    now_row = start_row
    now_column = start_column

    for idx, solve in enumerate(examples):
        
        for ch in solve:
            current_wbk.cell(row=now_row, column=now_column).value = ch
            now_column += 1

        for i in range(0, 4):
            current_wbk.cell(row=now_row, column=now_column+i).border = border

        if answers_list:
            for num in answers_list[idx]:
                current_wbk.cell(row=now_row, column=now_column).value = num
                now_column += 1

        now_row += 2
        now_column = start_column


def convert_xlsx_to_pdf(dir, xlsx_file):
    try:
        subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", xlsx_file, "--outdir", dir],
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    except Exception as e:
        print("Error:", e)


def create_examples(dir='', actions=['+', '-', '×', ':'], number_range=[1, 100], only_positive=True):

    m1 = generate_examples(types_action=actions, number_range=number_range, only_positive=only_positive)
    m2 = generate_examples(types_action=actions, number_range=number_range, only_positive=only_positive)

    print_examples_cell(wbk, m1[0], 3, 4)
    print_examples_cell(wbk, m2[0], 3, 17)

    print_examples_cell(wbk_answers, m1[0], 3, 4, m1[1])
    print_examples_cell(wbk_answers, m2[0], 3, 17, m2[1])

    wbObj.save(f'{dir}examples.xlsx')
    wbObj.close()

    wbObj_answers.save(f'{dir}examples_answers.xlsx')
    wbObj_answers.close()

    convert_xlsx_to_pdf(dir, f"{dir}examples.xlsx")
    convert_xlsx_to_pdf(dir, f"{dir}examples_answers.xlsx")

    os.remove(f'{dir}/examples.xlsx')
    os.remove(f'{dir}/examples_answers.xlsx')
