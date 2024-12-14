from PyQt5.QtWidgets import *

import openpyxl
from openpyxl.styles.borders import Border, Side

from random import randint
import random

import subprocess
import sys, os


class Window(QMainWindow):

    def __init__(self):
        super(Window, self).__init__()
        self.setWindowTitle("Генератор примеров")
        self.selected_folder = None

        style_border = 'medium'
        color_border = 'cccccc'
        self.border = Border(left=Side(style=style_border, color=color_border),
                        right=Side(style=style_border, color=color_border),
                        top=Side(style=style_border, color=color_border),
                        bottom=Side(style=style_border, color=color_border)
                        )

        self.setGeometry(100, 100, 600, 400)
        self.UiComponents()
        self.show()


    def UiComponents(self):
        self.themes_combo_box = QComboBox(self)
        self.themes_combo_box.setGeometry(10, 30, 250, 30)
        themes_list = ["Сложение целых чисел", "Не придумал", "Не придумал 2", "Не придумал 3"]
        self.themes_combo_box.addItems(themes_list)

        self.dif_combo_box = QComboBox(self)
        self.dif_combo_box.setGeometry(10, 100, 250, 30)
        dif_list = ["До 100", "До 200", "До 201", "До 202"]
        self.dif_combo_box.addItems(dif_list)

        label = QLabel(self)
        label.setText("Выбор темы")
        label.move(10, 0)

        label = QLabel(self)
        label.setText("Выбор сложности")
        label.move(10, 70)
        label.setFixedWidth(150)

        btn = QPushButton(self)
        btn.move(480, 350)
        btn.setText('Создать')
        btn.clicked.connect(self.create_examples)

        btn2 = QPushButton(self)
        btn2.move(10, 350)
        btn2.setText('Выбор папки')
        btn2.clicked.connect(self.select_folder)


    def select_folder(self):
        self.selected_folder = QFileDialog.getExistingDirectory(self, "Выберите папку")
        if self.selected_folder:
            print(f"Выбрана папка: {self.selected_folder}")


    def calc(self, action, n1, n2):
        match action:
            case '+':
                return n1 + n2
            case '-':
                return n1 - n2
            case ':':
                return n1 / n2
            case '×':
                return n1 * n2


    def generate_examples(self, n=22, types_action=['+', '-', '×', ':'], number_range=[1, 100], only_positive=False):
        examples_list = []
        result_list = []

        for _ in range(n):
            action = random.choice(types_action)

            num1 = randint(*number_range)
            num2 = randint(*number_range)

            if only_positive:
                num2 = randint(number_range[0], num1)

            num1_str, num2_str = list(str(num1)), list(str(num2))

            examples_list.append([*num1_str, action, *num2_str, '='])
            result_list.append(list(str(self.calc(action, num1, num2))))

        return [examples_list, result_list]


    def print_examples_cell(self, current_wbk, examples, start_row=4, start_column=5, answers_list=[]):

        now_row = start_row
        now_column = start_column

        for idx, solve in enumerate(examples):
            for ch in solve:
                current_wbk.cell(row=now_row, column=now_column).value = ch
                now_column += 1

            for i in range(0, 4):
                current_wbk.cell(row=now_row, column=now_column + i).border = self.border

            if answers_list:
                for num in answers_list[idx]:
                    current_wbk.cell(row=now_row, column=now_column).value = num
                    now_column += 1

            now_row += 2
            now_column = start_column


    def convert_xlsx_to_pdf(self, xlsx_file):
        try:
            subprocess.run(["libreoffice", "--headless", "--convert-to",
                            "pdf", xlsx_file, '--outdir', f'{self.selected_folder}/'])
            print("Done!")

        except Exception as e:
            print("Error:", e)


    def create_examples(self):
        if not self.selected_folder:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите папку!")
            return


        wbObj = openpyxl.load_workbook('list.xlsx')
        wbk = wbObj.worksheets[0]

        wbObj_answers = openpyxl.load_workbook('list.xlsx')
        wbk_answers = wbObj_answers.worksheets[0]

        m1 = self.generate_examples(types_action=['+', '-'], only_positive=True)
        m2 = self.generate_examples(types_action=['+', '-'], only_positive=True)

        self.print_examples_cell(wbk, m1[0], 3, 4)
        self.print_examples_cell(wbk, m2[0], 3, 17)

        self.print_examples_cell(wbk_answers, m1[0], 3, 4, m1[1])
        self.print_examples_cell(wbk_answers, m2[0], 3, 17, m2[1])

        wbObj.save(f'{self.selected_folder}/examples.xlsx')
        wbObj.close()

        wbObj_answers.save(f'{self.selected_folder}/examples_answers.xlsx')
        wbObj_answers.close()

        self.convert_xlsx_to_pdf(f'{self.selected_folder}/examples.xlsx')
        self.convert_xlsx_to_pdf(f'{self.selected_folder}/examples_answers.xlsx')

        os.remove(f'{self.selected_folder}/examples.xlsx')
        os.remove(f'{self.selected_folder}/examples_answers.xlsx')


App = QApplication(sys.argv)
window = Window()
sys.exit(App.exec())